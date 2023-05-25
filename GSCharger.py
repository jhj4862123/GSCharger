import os
import re
import shutil
from collections import defaultdict
from datetime import datetime
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox

import pandas as pd
import win32api
import win32com.client  # pip install pywin32
from PIL import ImageFile
from openpyxl import load_workbook
from openpyxl.styles import Color
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
from tqdm import tqdm

ImageFile.LOAD_TRUNCATED_IMAGES = True

############################ 이미지 폴더 선택 ########################################
root = Tk()
root.title("폴더 선택 창")  # 타이틀 설정

file_frame = Frame(root)
file_frame.pack(fill="x", padx=5, pady=5)

root.geometry("320x240")  # 가로 *세로 사이즈
root.resizable(False, False)  # 가로 *세로 사이즈 변경 가능 유무

dir_path = None  # 폴더 경로 담을 변수 생성

excelfilenum = 0


def folder_select():
    global dir_path
    dir_path = filedialog.askdirectory(initialdir="./", \
                                       title="폴더를 선택 해 주세요")  # folder 변수에 선택 폴더 경로 넣기
    if dir_path == '':
        messagebox.showwarning("경고", "폴더를 선택 하세요")  # 폴더 선택 안했을 때 메세지 출력
    else:
        res = os.listdir(dir_path)  # 폴더에 있는 파일 리스트 넣기
        if len(res) == 0:
            messagebox.showwarning("경고", "폴더내 파일이 없습니다.")
        else:
            root.destroy()


def set_value(cell_number, copy_value):
    wsMaster[cell_number] = str(copy_value)
    wsMaster[cell_number].fill = PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid')


def set_value2(cell_number, copy_value):
    wsNew[cell_number] = str(copy_value)
    wsNew[cell_number].alignment = Alignment(horizontal='center', vertical='center')
    wsNew[cell_number].fill = PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid')


def warning():
    if len(warningphoto) != 0:
        f = open(f"생성되지 않은 파일_{ss}.txt", 'w')
        f.write(ss + " 기준\n")
        f.write('#' * 40 + '\n')

        print(len(warningphoto))

        for count in range(len(warningphoto)): # 사진이 없는 경우
            if warningphoto[count] != 0:
                f.write('충전기 [ {0} ]번 파일은 그림이 없어 생성 실패했습니다.\n'.format(warningphoto))  # 가운데 정렬

        for cnt in range(len(noChargNum)): # 충전기 번호가 없는 경우
            if noChargNum[cnt] != 0:
                f.write('충전기 [ {0} ]번 파일은 등록되지 않은 충전기 번호이니 참고하시기 바랍니다.\n'.format(noChargNum))  # 가운데 정렬
            # f.write('충전기 [ {0:>10} ]이 파일 생성에 실패하였습니다.\n'.format(warningphoto[i]))  # 오른쪽 정렬
        f.write('#' * 40)
        f.close()

        f = open(f"생성되지 않은 파일_{ss}.txt", 'r')
        data = f.read()
        f.close()
        win32api.MessageBox(0, f"{data}", "파일생성오류", 16)
        print(data)
        shutil.move(f"생성되지 않은 파일_{ss}.txt", finishpath + f"생성되지 않은 파일_{ss}.txt")


btn_active_dir = Button(file_frame, text="충전기 사진을 선택해 주세요. \n\n사진 형식 : 충전기번호_1.jpg\n ex) 1234_1.jpg", \
                        font=36, width=24, padx=10, pady=20, command=folder_select)
btn_active_dir.pack(padx=5, pady=5)

root.mainloop()

############################ 경로 및 양식 ########################################
now = datetime.now()
s = now.strftime("%Y-%m-%d")
ss = now.strftime("%Y-%m-%d %H시 %M분")
finishpath = '완료폴더/'
newpath = finishpath + s

photosrc = dir_path + '/'
movephoto = newpath + '/완료된 사진/'
move_resize_photo = newpath + '/축소 사진/'
resultpath = newpath + '/결과/'
movefilesrc = '완료폴더/'
path = './점검데이터.xlsx'
pwd = os.getcwd()

j = 1

if not os.path.exists(newpath):
    os.makedirs(newpath)

if not os.path.exists(movephoto):
    os.makedirs(movephoto)

if not os.path.exists(move_resize_photo):
    os.makedirs(move_resize_photo)

if not os.path.exists(resultpath):
    os.makedirs(resultpath)

data = pd.read_excel('점검데이터.xlsx', sheet_name='점검정보')

base = photosrc
print("\nbase : ", base)
count_photo = []  # 사진의 갯수

################## vlookup #################
df = pd.read_excel('점검데이터.xlsx')
wbSlave = load_workbook('점검데이터.xlsx')
wsSlave = wbSlave['점검정보']

for i in range(1, len(df.loc[1])):
    column_chr = get_column_letter(i + 1)

    wsSlave[40][i].value = "=VLOOKUP(" + column_chr + "2,기준정보!$B:$AM,3,FALSE)"
    wsSlave[41][i].value = "=VLOOKUP(" + column_chr + "2,기준정보!$B:$AM,11,FALSE)"
    wsSlave[42][i].value = "=VLOOKUP(" + column_chr + "2,기준정보!$B:$AM,37,FALSE) & \"/\" " + \
                           "& VLOOKUP(" + column_chr + "2,기준정보!$B:$AM,38,FALSE)"
    wsSlave[43][i].value = "=VLOOKUP(" + column_chr + "2,기준정보!$B:$AM,30,FALSE)"
    wsSlave[44][i].value = "=VLOOKUP(" + column_chr + "2,기준정보!$B:$AM,34,FALSE)"
    wsSlave[45][i].value = "=VLOOKUP(" + column_chr + "2,기준정보!$B:$AM,21,FALSE)"

    if wsSlave[40][i].value == "#N/A":
        fill = PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid')
        wsSlave[40][i].fill = fill

wbSlave.save("점검데이터.xlsx")
# wbSlave.close()

############## Win32 ##############
# excel 사용할 수 있게 설정
excel = win32com.client.Dispatch("Excel.Application")
# 임시 Workbook 객체 생성 및 엑셀 열기
temp_wb = excel.Workbooks.Open(pwd + "\점검데이터.xlsx")
# 저장
temp_wb.Save()
# excel 종료
temp_wb.Close()
############################ 파일 분리 ############################################################
file_names = []

file_names = os.listdir(dir_path)

for name in file_names:
    src_name = name
    temp_name = re.split('[,|_|.|-]', name)
    src = os.path.join(photosrc, name)

    if len(temp_name) != 3:

        for j in range(0, len(temp_name) - 2):
            print(f"글자 분리 : {temp_name[j]}")
            print(f"src : {src}")
            dst = temp_name[j] + '-' + temp_name[-2] + '.jpg'
            dst = os.path.join(photosrc, dst)
            print(f"dst : {dst}")
            shutil.copyfile(src, dst)
        shutil.move(os.path.join(dir_path, src_name), os.path.join(movephoto, src_name))

    else:
        dst = temp_name[0] + '-' + temp_name[-2] + '.jpg'
        dst = os.path.join(photosrc, dst)
        print(f"dst : {dst}")
        shutil.copyfile(src, dst)
        shutil.move(os.path.join(dir_path, src_name), os.path.join(movephoto, src_name))

############################# 충전기 갯수 카운트 및 이미지 리사이즈 ########################################
from PIL import Image

chargernum = 1  # 충전기의 갯수
사진없는개수 = {}
사진없는개수 = defaultdict(int)  # 사진이 없는 개수를 기록할 사전

for name in data.iloc[0, 1:]:  # None 없애기
    chargernum += 1
    for j in range(1, 7):
        fileName = os.path.join(base, str(name) + "_" + str(j) + ".jpg")
        tempName = os.path.join(base, str(name) + "-" + str(j) + ".jpg")

        if os.path.exists(fileName):
            img = Image.open(fileName)

        elif os.path.exists(tempName):
            shutil.move(tempName, fileName)
            img = Image.open(fileName)
        else:
            사진없는개수[name] += 1  # 사진이 없는 개수를 증가시킴
            continue
        img = img.convert('RGB')

        width, height = img.size[:2]

        if height >= width:
            img = img.resize((277, 277))
            resize_img = img.save(base + str(name) + "-" + str(j) + "(resize).jpg")

        else:
            img = img.resize((312, 277))
            resize_img = img.save(base + str(name) + "-" + str(j) + "(resize).jpg")

############################# 파일 생성 실패 시 배경색 채우기 ########################################
data = pd.read_excel('점검데이터.xlsx', sheet_name='점검정보')
count = 0

wbSlave = load_workbook('점검데이터.xlsx', data_only=True)
wsSlave = wbSlave['점검정보']

for name in data.iloc[0]:  # None 없애기
    count += 1
    if 사진없는개수.get(name) == 6:
        temp = get_column_letter(count)
        cell_number = temp + '2'
        wsSlave[cell_number].fill = PatternFill(fill_type='solid', fgColor=Color('FF0000'))

wbSlave.save("점검데이터.xlsx")

############################# 양식 ########################################

wbMaster = load_workbook('정기점검보고서.xlsx')
wsMaster = wbMaster['정기점검보고서']
wsNew = wbMaster['정기점검보고서']

wbSlave = load_workbook('점검데이터.xlsx', data_only=True)
wsSlave = wbSlave['점검정보']

############################# 추가해서 수정해야함 ########################################

warningphoto = []
noChargNum = []

existphoto = []
############################# 변수들 ########################################
for i in tqdm(range(chargernum - 1)):

    충전기번호 = wsSlave.cell(row=2, column=i + 2).value
    print(사진없는개수.get(충전기번호))
    if 사진없는개수.get(충전기번호) == 6:
        # for x in 사진없는개수:
        warningphoto.append(충전기번호)
        #continue

    wbMaster = load_workbook('정기점검보고서.xlsx')
    wsMaster = wbMaster['정기점검보고서']
    wsNew = wbMaster['정기점검보고서']
    wbSlave = load_workbook('점검데이터.xlsx', data_only=True)
    # slavestandard = wbSlave['참조데이터']
    wsSlave = wbSlave['점검정보']

    set_value('G7', 충전기번호)

    점검일자 = wsSlave['4'][1 + i].value
    set_value('C5', 점검일자)

    if 점검일자 == "#N/A":
        print("점검일자가 설정되지 않았습니다.")
    else:
        try:
            day001 = pd.to_datetime(점검일자, format='%Y-%m-%d')
            day001 = day001.date()
            print(day001)
        except ValueError:
            print("유효한 날짜 형식이 아닙니다.")

    점검자 = wsSlave['3'][1 + i].value
    set_value('H5', 점검자)

    주변온도 = wsSlave['6'][1 + i].value
    set_value('C6', 주변온도)

    습도 = wsSlave['7'][1 + i].value
    set_value('H6', 습도)

    충전소이름 = str(wsSlave['40'][1 + i].value)
    set_value('B8', "충전소 이름 : " + 충전소이름)

    if 충전소이름 == "#N/A":
        noChargNum.append(충전기번호)
        print(str(충전기번호) + "는 등록되지 않은 충전기입니다.")

    continue

    충전기제조사 = str(wsSlave['41'][1 + i].value)
    set_value('B9', "충전기 제조사 : " + 충전기제조사)

    충전소좌표 = str(wsSlave['42'][1 + i].value)
    set_value('B10', "충전소 좌표 : " + 충전소좌표)

    충전소주소 = str(wsSlave['43'][1 + i].value)
    set_value('B11', "충전소 주소 : " + 충전소주소)

    충전기위치 = str(wsSlave['44'][1 + i].value)
    set_value('B12', "충전기 위치 : " + 충전기위치)

    전압 = str(wsSlave['11'][1 + i].value)
    set_value('C15', 전압 + "V")

    충전기용량 = str(wsSlave['45'][1 + i].value)
    set_value('E15', 충전기용량)

    수량 = str(wsSlave['5'][1 + i].value)
    set_value('G9', "충전기 총수량 : " + 수량 + "대")

    무료주차 = wsSlave['27'][1 + i].value
    set_value('D23', 무료주차)

    급속완속 = wsSlave['28'][1 + i].value
    set_value('J22', 급속완속)

    적정조명 = wsSlave['12'][1 + i].value
    set_value('G33', 적정조명)

    비상정지 = wsSlave['18'][1 + i].value
    set_value('G37', 비상정지)

    환경부카드 = wsSlave['29'][1 + i].value
    set_value('J37', 환경부카드)

    커넥터균열 = wsSlave['30'][1 + i].value
    set_value('G42', 커넥터균열)

    볼라드고정 = wsSlave['31'][1 + i].value
    set_value('D43', 볼라드고정)

    보호잠금장치 = wsSlave['32'][1 + i].value
    set_value('G43', 보호잠금장치)

    위험표지 = wsSlave['33'][1 + i].value
    set_value('D51', 위험표지)

    분전함lock = wsSlave['32'][1 + i].value
    set_value('D52', 분전함lock)

    분전함접지 = wsSlave['19'][1 + i].value
    set_value('D53', 분전함접지)

    과전류차단기 = wsSlave['15'][1 + i].value
    set_value('J49', 과전류차단기)

    누전차단기 = wsSlave['16'][1 + i].value
    set_value('J50', 누전차단기)

    감도전류 = wsSlave['17'][1 + i].value
    set_value('J51', 감도전류)

    전선굵기 = wsSlave['34'][1 + i].value
    set_value('J52', 전선굵기)

    접지저항 = wsSlave['13'][1 + i].value
    set_value('E58', 접지저항)

    접지선굵기 = wsSlave['35'][1 + i].value
    set_value('E59', 접지선굵기)

    소화설비 = wsSlave['25'][1 + i].value
    set_value('D67', 소화설비)

    passes = ['D21', 'D22', 'D23', 'D27', 'D28', 'D29', 'D34', 'D35', 'D50', 'D51', 'D54', 'G21', 'G22', 'G27', 'G28',
              'G29', 'G49', 'G50', 'G51', 'J21', 'J27', 'J28', 'J33', 'J36']

    for cell in passes:
        wsMaster[cell] = 'Y'

    ############################ 사진 ########################################

    src_img = []
    # 사진이 하나도 없으면 이 For문을 돌지 않게
    for j in range(0, 6):
        path1 = os.path.join(base, str(충전기번호) + "_" + str(j + 1) + ".jpg")
        src_img.append(path1)

        if os.path.exists(src_img[j]):
            fileName = os.path.join(base,
                                    str(충전기번호) + "-" + str(j + 1) + "(resize).jpg")  # 여기가 없음!!!!!!!!!!!!!!!!!!!!!!!!!!
            if os.path.exists(fileName):
                img1 = Image(fileName)
                print(fileName)
                if (j == 0):
                    position = 'B90'
                elif (j == 1):
                    position = 'G90'
                elif (j == 2):
                    position = 'B103'
                elif (j == 3):
                    position = 'G103'
                elif (j == 4):
                    position = 'B116'
                elif (j == 5):
                    position = 'G116'
                else:
                    pass
                wsMaster.add_image(img1, position)
            # shutil.move(photosrc + str(충전기번호) + "_" + str(j + 1) + ".jpg", movephoto + str(충전기번호) + "-" + str(j + 1) + ".jpg")

            # else:
            #    print(f"{fileName} (===>fileName) ")
        # else:
        #   print(f"{src_img[j]} (src image file) 사진이 없습니다.")
    ############################# 출력형식 ########################################

    if 사진없는개수[충전기번호] != 0 or wsSlave[40][i].value != "#N/A":  # 사진이 있거나 충전기 번호가 있는 경우
        wbMaster.save(str(충전기번호) + "-" + str(점검자) + "-" + str(day001) + ".xlsx")
    else:
        continue
    shutil.move(str(충전기번호) + "-" + str(점검자) + "-" + str(day001) + ".xlsx",
                resultpath + "/" + str(충전기번호) + "-" + str(점검자) + "-" + str(day001) + ".xlsx")
    excelfilenum = excelfilenum + 1

    wbMaster.close()
    print("\n" + str(충전기번호) + "-" + str(점검자) + "-" + str(day001) + ".xlsx" + " 파일이 생성되었습니다.")

shutil.copy(path, newpath + '\점검데이터(' + ss + ').xlsx')

file_list = os.listdir(base)  # 폴더안의 파일 리스트를 얻습니다.

for item in file_list:
    if item[-12:] == "(resize).jpg":  # item[-12:] 마지막 12글자
        shutil.move(photosrc + item, move_resize_photo + item)

files = os.listdir(photosrc)

for f in files:
    shutil.move(photosrc + f, movephoto + f)

print("총 " + str(excelfilenum) + "개의 파일이 생성되었습니다.")
warning()
# shutil.move("생성되지 않은 파일.txt", newpath + '\생성되지 않은 파일(' + ss +').txt')


input("엔터를 누르면 종료됩니다.")
exit()
